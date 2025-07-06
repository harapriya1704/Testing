from openpyxl import load_workbook
from logger import logger

def read_excel_with_required_columns(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
        required_columns = [
            "Fiscal Week", "Date", "Order Number", 
            "Improve Text", "Glassbox Link", "Sat/Dissat"
        ]
        
        missing = [col for col in required_columns if col not in headers]
        if missing:
            logger.error(f"Missing columns in {file_path}: {', '.join(missing)}")
            return []
        
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            order_number = row[headers["Order Number"]]
            sat_dissat = row[headers["Sat/Dissat"]]

            if (str(order_number).strip().lower() not in ["", "null", "none"] and str(sat_dissat).strip().upper() == "DSAT"):
                data.append({
                    "fiscal_week": row[headers["Fiscal Week"]],
                    "date": row[headers["Date"]],
                    "order_number": row[headers["Order Number"]],
                    "improve_text": row[headers["Improve Text"]],
                    "glassbox_link": row[headers["Glassbox Link"]],
                    "sat_dissat": "DSAT"
                })
                
        return data
        
    except Exception as e:
        logger.error(f"Error reading {file_path}: {e}", exc_info=True)
        return []
