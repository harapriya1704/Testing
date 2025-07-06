import json
from openpyxl import Workbook, load_workbook
import os

def initialize_output_excel(output_path):
    wb = Workbook()
    ws = wb.active
    headers = [
        "Fiscal Week", "Date", "Order Number", "Sat/Dissat", "Improve Text",
        "GIA Insights", "Global_DellCEMSessionCookie_CSH", "Global_MCMID_CSH",
        "Client-Sessions", "Server-Sessions", "Order Details"
    ]
    ws.append(headers)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✅ Initialized output Excel at {output_path}")

def append_session_to_excel(entry, output_path):
    if not os.path.exists(output_path):
        initialize_output_excel(output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    row = [
        entry.get("fiscal_week", ""),
        entry.get("date", ""),
        entry.get("order_number", ""),
        entry.get("sat_dissat", ""),
        entry.get("improve_text", ""),
        entry.get("gia_insights", ""),
        entry.get("Global_DellCEMSessionCookie_CSH", ""),
        entry.get("Global_MCMID_CSH", ""),
        entry.get("Client-Sessions", ""),
        entry.get("Server-Sessions", ""),
        json.dumps(entry.get("order_details", {}))
    ]
    ws.append(row)
    wb.save(output_path)
    print(f"📥 Appended session for order {entry.get('order_number')} to Excel.")

def update_last_row_with_order_details(order_details, output_path):
    wb = load_workbook(output_path)
    ws = wb.active
    last_row = ws.max_row
    ws.cell(row=last_row, column=11).value = json.dumps(order_details)
    wb.save(output_path)
    print(f"🔄 Updated order details for last row in Excel.")
